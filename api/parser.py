"""
Parse sbr-stats.xlsx into Python dicts cached at startup.
"""
from __future__ import annotations

import os
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

XLSX_PATH = os.environ.get("SBR_XLSX_PATH", os.path.join(os.path.dirname(__file__), "..", "sbr-stats.xlsx"))

# Seasons that store finishing positions (not points) in race columns
POSITION_SEASONS = {"S1", "S2", "S3", "S4", "S5", "S6", "WEC"}

# Sheets that are not season data
NON_SEASON_SHEETS = {"Less Ugly", "Wins", "Podiums", "Participation"}

# Map sheet names → canonical display names (identity unless overridden)
SHEET_DISPLAY_NAMES: Dict[str, str] = {
    "SUPER GT 22": "Super GT '22",
    "GT4 America 22": "GT4 America '22",
    "GT4 Australia": "GT4 Australia",
}


def _normalize_cell(val: Any) -> Any:
    """Convert raw cell value to a clean Python value."""
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip().rstrip()
        if stripped in ("DNP",):
            return "DNP"
        if stripped in ("DNS",):
            return "DNS"
        if stripped in ("DNF",):
            return "DNF"
        if stripped in ("-", "-  ", "- ", "–"):
            return "DNS"
        # Asterisked points e.g. "48*"
        if re.match(r"^\d+\*$", stripped):
            return int(stripped[:-1])
        # Try int/float
        try:
            return int(stripped)
        except ValueError:
            pass
        try:
            return float(stripped)
        except ValueError:
            pass
        return stripped if stripped else None
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


def _is_double_format(headers: list[str]) -> bool:
    """Detect seasons where each round has two races (feature + reverse grid)."""
    has_r = any(re.match(r"^R\d+$", h) for h in headers if h)
    has_col = any(re.match(r"^Column\d+$", h) for h in headers if h)
    return has_r and has_col


def _race_headers(headers: list[str]) -> list[tuple[str, str]]:
    """
    Return list of (kind, label) for each race column.
    kind is 'feature', 'reverse', or 'single'.
    Handles both single-format (R1, R2…) and double-format (R1, Column1, R2, Column2…).
    Also handles the F3 oddity where Column1 appears before R1.
    """
    result = []
    for h in headers:
        if h is None:
            continue
        if re.match(r"^R\d+$", h):
            result.append(("feature", h))
        elif re.match(r"^Column\d+$", h):
            result.append(("reverse", h))
    return result


def _parse_season_sheet(ws) -> dict:
    """Parse a single season worksheet into a structured dict."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header_row = [str(h).strip() if h is not None else None for h in rows[0]]

    # Detect race format
    double = _is_double_format(header_row)
    race_format = "double" if double else "single"

    # Find column indices for fixed fields
    def col(name: str) -> Optional[int]:
        for i, h in enumerate(header_row):
            if h and h.upper() == name.upper():
                return i
        return None

    # Support both "DRIVER" and "DRIVERS"
    driver_col = col("DRIVER") if col("DRIVER") is not None else col("DRIVERS")
    pos_col = col("POS") if col("POS") is not None else col("RANK")
    wins_col = col("WINS")
    podiums_col = col("PODIUMS")
    dns_col = col("DNS")
    total_col = col("TOTAL") if col("TOTAL") is not None else col("TOTAl")

    # Race column indices: list of (kind, col_index, round_number)
    race_cols: List[Tuple[str, int, int]] = []
    round_num = 0
    for i, h in enumerate(header_row):
        if h is None:
            continue
        if re.match(r"^R\d+$", h):
            round_num += 1
            kind = "feature" if double else "single"
            race_cols.append((kind, i, round_num))
        elif re.match(r"^Column\d+$", h):
            # reverse grid — same round as the preceding R
            race_cols.append(("reverse", i, round_num))

    # Build standings
    standings = []
    for row in rows[1:]:
        if driver_col is None or driver_col >= len(row):
            continue
        driver = _normalize_cell(row[driver_col])
        if not driver or driver == "DNP":
            continue

        pos_val = _normalize_cell(row[pos_col]) if pos_col is not None else None
        wins_val = _normalize_cell(row[wins_col]) if wins_col is not None else 0
        podiums_val = _normalize_cell(row[podiums_col]) if podiums_col is not None else 0
        dns_val = _normalize_cell(row[dns_col]) if dns_col is not None else None
        total_val = _normalize_cell(row[total_col]) if total_col is not None else None

        # Build race results
        if double:
            # Group into rounds: {round_num: {"feature": val, "reverse": val}}
            rounds_dict: Dict[int, dict] = {}
            for kind, ci, rnum in race_cols:
                val = _normalize_cell(row[ci]) if ci < len(row) else None
                if rnum not in rounds_dict:
                    rounds_dict[rnum] = {"feature": None, "reverse": None}
                rounds_dict[rnum][kind] = val
            rounds = [rounds_dict[r] for r in sorted(rounds_dict)]
        else:
            rounds = []
            for kind, ci, rnum in race_cols:
                val = _normalize_cell(row[ci]) if ci < len(row) else None
                rounds.append({"result": val})

        standings.append({
            "pos": pos_val,
            "driver": str(driver),
            "wins": wins_val if isinstance(wins_val, int) else 0,
            "podiums": podiums_val if isinstance(podiums_val, int) else 0,
            "dns": dns_val if isinstance(dns_val, int) else 0,
            "total": total_val,
            "rounds": rounds,
        })

    # Sort standings by pos (None last)
    standings.sort(key=lambda s: (s["pos"] is None, s["pos"] if isinstance(s["pos"], int) else 9999))

    champion = standings[0]["driver"] if standings else None
    num_rounds = max((len(s["rounds"]) for s in standings), default=0)

    # Compute race winners per round
    race_winners = _compute_race_winners(standings, double, num_rounds)

    return {
        "race_format": race_format,
        "score_type": "position",  # will be overridden in load()
        "num_rounds": num_rounds,
        "champion": champion,
        "standings": standings,
        "race_winners": race_winners,
    }


def _compute_race_winners(standings: List[dict], double: bool, num_rounds: int) -> List[dict]:
    """For each round, find which driver had the best result."""
    winners = []
    for r in range(num_rounds):
        if double:
            feature_winner = _round_winner(standings, r, "feature")
            reverse_winner = _round_winner(standings, r, "reverse")
            winners.append({
                "round": r + 1,
                "feature_winner": feature_winner,
                "reverse_winner": reverse_winner,
            })
        else:
            winner = _round_winner(standings, r, "result")
            winners.append({"round": r + 1, "winner": winner})
    return winners


def _round_winner(standings: List[dict], round_idx: int, key: str) -> Optional[str]:
    """
    Find the winner of a single race.
    - For position seasons: lowest numeric result wins (1st place).
    - For points seasons: highest numeric result wins.
    We return the driver with the best numeric result; skip DNS/DNF/None.
    """
    best_driver = None
    best_val = None
    # We'll determine sort direction per the caller context (handled outside, this func
    # just returns the driver with the numeric extremum)
    # Convention: if values look like positions (small ints <=20), lowest wins;
    # if large ints (points), highest wins.
    candidates = []
    for s in standings:
        if round_idx >= len(s["rounds"]):
            continue
        val = s["rounds"][round_idx].get(key)
        if isinstance(val, int):
            candidates.append((val, s["driver"]))

    if not candidates:
        return None

    # Heuristic: if max value <= 30, treat as positions (lower = better)
    max_val = max(v for v, _ in candidates)
    if max_val <= 30:
        # position mode: lowest wins
        candidates.sort(key=lambda x: x[0])
    else:
        # points mode: highest wins
        candidates.sort(key=lambda x: -x[0])

    return candidates[0][1] if candidates else None


def _parse_wins_sheet(ws) -> Dict[str, Dict[str, Any]]:
    """Parse the Wins collector sheet → {driver: {season: count}}"""
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    result = {}
    for row in rows[1:]:
        driver = _normalize_cell(row[0])
        if not driver or not isinstance(driver, str):
            continue
        total = _normalize_cell(row[1])
        per_season = {}
        for i, h in enumerate(headers[2:], start=2):
            if h:
                val = _normalize_cell(row[i]) if i < len(row) else None
                per_season[h] = val
        result[driver] = {"total": total, "per_season": per_season}
    return result


def _parse_podiums_sheet(ws) -> dict[str, dict[str, Any]]:
    return _parse_wins_sheet(ws)  # identical structure


def _parse_participation_sheet(ws) -> Dict[str, Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    result = {}
    for row in rows[1:]:
        driver = _normalize_cell(row[0])
        if not driver or not isinstance(driver, str):
            continue
        dnp_count = _normalize_cell(row[1])
        per_season = {}
        for i, h in enumerate(headers[2:], start=2):
            if h:
                val = _normalize_cell(row[i]) if i < len(row) else None
                per_season[h] = val
        result[driver] = {"dnps": dnp_count, "per_season": per_season}
    return result


def _parse_less_ugly(ws) -> dict:
    """Parse the Less Ugly summary sheet for championships."""
    rows = list(ws.iter_rows(values_only=True))
    # Championships are in columns 20-23 (0-indexed) based on observed layout
    championships: dict[str, int] = {}
    season_champs: list[dict] = []

    for row in rows[1:]:
        # Driver championships: col 8 = driver, col 9 = count
        if len(row) > 9:
            drv = _normalize_cell(row[8])
            cnt = _normalize_cell(row[9])
            if drv and isinstance(drv, str) and isinstance(cnt, int):
                championships[drv] = cnt

        # Season champions: col 20 = season name, col 21 = driver champ, col 22 = team champ
        if len(row) > 21:
            season = _normalize_cell(row[20])
            champ = _normalize_cell(row[21])
            if season and isinstance(season, str) and champ:
                season_champs.append({"season": season, "champion": champ})

    return {"championships": championships, "season_champions": season_champs}


_cache: dict = {}


def load() -> dict:
    """Load and cache all data from the spreadsheet. Returns the cache dict."""
    global _cache
    if _cache:
        return _cache

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    wins_data = _parse_wins_sheet(wb["Wins"])
    podiums_data = _parse_podiums_sheet(wb["Podiums"])
    participation_data = _parse_participation_sheet(wb["Participation"])
    less_ugly = _parse_less_ugly(wb["Less Ugly"])

    seasons: dict[str, dict] = {}
    season_order: list[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in NON_SEASON_SHEETS:
            continue
        ws = wb[sheet_name]
        parsed = _parse_season_sheet(ws)
        if not parsed:
            continue

        display_name = SHEET_DISPLAY_NAMES.get(sheet_name, sheet_name)
        parsed["name"] = sheet_name
        parsed["display_name"] = display_name
        parsed["score_type"] = "position" if sheet_name in POSITION_SEASONS else "points"

        # Attach wins/podiums per driver from collector sheets
        for standing in parsed["standings"]:
            drv = standing["driver"]
            w = wins_data.get(drv, {})
            p = podiums_data.get(drv, {})
            # Prefer data from the season sheet itself (already in standing["wins"/"podiums"])
            # but fall back to collector if missing
            if standing["wins"] == 0 and sheet_name in w.get("per_season", {}):
                ws_wins = w["per_season"][sheet_name]
                if isinstance(ws_wins, int):
                    standing["wins"] = ws_wins
            if standing["podiums"] == 0 and sheet_name in p.get("per_season", {}):
                ws_pods = p["per_season"][sheet_name]
                if isinstance(ws_pods, int):
                    standing["podiums"] = ws_pods

        seasons[sheet_name] = parsed
        season_order.append(sheet_name)

    # Build driver index
    all_drivers: set[str] = set()
    for drv in wins_data:
        all_drivers.add(drv)
    for drv in podiums_data:
        all_drivers.add(drv)
    for season in seasons.values():
        for s in season["standings"]:
            all_drivers.add(s["driver"])

    _cache = {
        "wins": wins_data,
        "podiums": podiums_data,
        "participation": participation_data,
        "less_ugly": less_ugly,
        "seasons": seasons,
        "season_order": season_order,
        "all_drivers": sorted(all_drivers),
    }
    return _cache

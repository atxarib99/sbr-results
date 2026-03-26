"""
scripts/import_xlsx.py

Idempotent xlsx → MariaDB importer.

Truncates and reloads DATA tables only. Config tables (canonical_players,
driver_aliases, classes, points_structures, points_structure_entries,
season_points_structure) are never touched.

Drivers are upserted (INSERT ... ON DUPLICATE KEY UPDATE) so their IDs
remain stable across re-imports, preserving driver_aliases FK references.

Usage:
    python scripts/import_xlsx.py [--xlsx path/to/sbr-stats.xlsx]

Environment variables (all optional):
    DB_HOST        default: localhost
    DB_PORT        default: 3306
    DB_USER        default: sbr
    DB_PASSWORD    default: sbr
    DB_NAME        default: sbr
    SBR_XLSX_PATH  default: sbr-stats.xlsx (relative to project root)
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

import pymysql
import pymysql.cursors
import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = os.getenv("SBR_XLSX_PATH", str(PROJECT_ROOT / "sbr-stats.xlsx"))

# ── DB connection ─────────────────────────────────────────────────────────────

DB_CONFIG: dict = {
    "host":        os.getenv("DB_HOST", "localhost"),
    "port":        int(os.getenv("DB_PORT", "3306")),
    "user":        os.getenv("DB_USER", "sbr"),
    "password":    os.getenv("DB_PASSWORD", "sbr"),
    "database":    os.getenv("DB_NAME", "sbr"),
    "charset":     "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
    "autocommit":  False,
}

# ── Season metadata ───────────────────────────────────────────────────────────

# Sheets that store finishing positions (1st, 2nd…) rather than points.
POSITION_SEASONS: set[str] = {"S1", "S2", "S3", "S4", "S5", "S6", "WEC"}

# Sheets that are not season data — skip entirely.
NON_SEASON_SHEETS: set[str] = {"Less Ugly", "Wins", "Podiums", "Participation"}

# Optional display name overrides.
SHEET_DISPLAY_NAMES: dict[str, str] = {
    "SUPER GT 22":   "Super GT '22",
    "GT4 America 22": "GT4 America '22",
    "GT4 Australia": "GT4 Australia",
}

# Per-season config flags.
# Edit this dict as new seasons are added.
# has_drop_round: worst round excluded from championship total
# is_multiclass:  multiple classes run simultaneously (assign classes via seed_config.sql)
SEASON_FLAGS: dict[str, dict] = {
    # "IMSA": {"is_multiclass": True},
    # "SomeSeasonName": {"has_drop_round": True},
}

# ── Cell normalisation ────────────────────────────────────────────────────────

def normalize_cell(val) -> tuple[float | None, str | None, bool]:
    """
    Decompose a raw xlsx cell value into (value_numeric, value_flag, is_asterisked).

    value_numeric : float | None — the numeric value if applicable
    value_flag    : 'dns' | 'dnf' | 'dnp' | None
    is_asterisked : bool — True if the original value had a trailing *
    """
    if val is None:
        return None, None, False
    if isinstance(val, (int, float)):
        return float(val), None, False
    if isinstance(val, str):
        s = val.strip()
        upper = s.upper()
        if upper == "DNP":
            return None, "dnp", False
        if upper == "DNF":
            return None, "dnf", False
        if upper in ("DNS", "-", "–", "- ", "-  "):
            return None, "dns", False
        # Asterisked number e.g. "48*"
        m = re.match(r"^(\d+(?:\.\d+)?)\*$", s)
        if m:
            return float(m.group(1)), None, True
        try:
            return float(int(s)), None, False
        except ValueError:
            pass
        try:
            return float(s), None, False
        except ValueError:
            pass
    return None, None, False


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def is_double_format(headers: list) -> bool:
    """True if headers contain both R\d+ and Column\d+ patterns."""
    has_r   = any(re.match(r"^R\d+$",      h) for h in headers if h)
    has_col = any(re.match(r"^Column\d+$", h) for h in headers if h)
    return has_r and has_col


def find_col(headers: list, *names: str) -> int | None:
    """Return index of first header matching any of names (case-insensitive)."""
    for name in names:
        for i, h in enumerate(headers):
            if h and h.upper() == name.upper():
                return i
    return None


def parse_race_columns(headers: list, double: bool) -> list[tuple[str, int, int]]:
    """
    Returns list of (sub_type, col_index, round_number).
    sub_type: 'single' | 'feature' | 'reverse'
    """
    result = []
    round_num = 0
    for i, h in enumerate(headers):
        if h is None:
            continue
        if re.match(r"^R\d+$", h):
            round_num += 1
            result.append(("feature" if double else "single", i, round_num))
        elif re.match(r"^Column\d+$", h):
            result.append(("reverse", i, round_num))
    return result


# ── Main import ───────────────────────────────────────────────────────────────

def run_import(xlsx_path: str) -> None:
    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── Pre-scan all season sheets ────────────────────────────────────────────
    sheet_data: list[dict] = []
    all_raw_names: set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name in NON_SEASON_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h is not None else None for h in rows[0]]
        driver_col = find_col(headers, "DRIVER", "DRIVERS")
        if driver_col is None:
            print(f"  [skip] {sheet_name!r}: no DRIVER column found")
            continue

        pos_col     = find_col(headers, "POS", "RANK")
        double      = is_double_format(headers)
        race_cols   = parse_race_columns(headers, double)

        # Collect driver names and find champion (POS=1)
        champion_raw: str | None = None
        sheet_drivers: list[tuple[str, int | None]] = []  # (raw_name, pos)

        for row in rows[1:]:
            if driver_col >= len(row):
                continue
            raw = row[driver_col]
            if raw is None:
                continue
            raw_str = str(raw).strip()
            if not raw_str or raw_str.upper() == "DNP":
                continue

            pos_val = None
            if pos_col is not None and pos_col < len(row):
                num, flag, _ = normalize_cell(row[pos_col])
                pos_val = int(num) if num is not None else None

            all_raw_names.add(raw_str)
            sheet_drivers.append((raw_str, pos_val))

        # Champion = driver with pos=1
        pos1 = [name for name, pos in sheet_drivers if pos == 1]
        champion_raw = pos1[0] if pos1 else (sheet_drivers[0][0] if sheet_drivers else None)

        sheet_data.append({
            "sheet_name":    sheet_name,
            "rows":          rows,
            "headers":       headers,
            "driver_col":    driver_col,
            "race_cols":     race_cols,
            "double":        double,
            "champion_raw":  champion_raw,
        })

    print(f"  Found {len(sheet_data)} season sheets, {len(all_raw_names)} unique driver names")

    # ── Connect and import ────────────────────────────────────────────────────
    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:

            # Truncate data tables (disable FK checks so order doesn't matter)
            print("Truncating data tables...")
            cur.execute("SET foreign_key_checks = 0")
            for tbl in ("race_results", "driver_season_class", "rounds", "seasons"):
                cur.execute(f"TRUNCATE TABLE `{tbl}`")
            cur.execute("SET foreign_key_checks = 1")

            # Upsert drivers — stable IDs preserve driver_aliases FKs
            print(f"Upserting {len(all_raw_names)} drivers...")
            for raw_name in sorted(all_raw_names):
                cur.execute(
                    "INSERT INTO drivers (raw_name) VALUES (%s) "
                    "ON DUPLICATE KEY UPDATE raw_name = raw_name",
                    (raw_name,),
                )

            # Build raw_name → id lookup
            cur.execute("SELECT id, raw_name FROM drivers")
            driver_id_map: dict[str, int] = {
                row["raw_name"]: row["id"] for row in cur.fetchall()
            }

            # Insert seasons, rounds, race_results
            total_results = 0
            for sort_order, sd in enumerate(sheet_data):
                sheet_name  = sd["sheet_name"]
                rows        = sd["rows"]
                headers     = sd["headers"]
                driver_col  = sd["driver_col"]
                race_cols   = sd["race_cols"]
                double      = sd["double"]
                champion_raw = sd["champion_raw"]

                display_name = SHEET_DISPLAY_NAMES.get(sheet_name, sheet_name)
                score_type   = "position" if sheet_name in POSITION_SEASONS else "points"
                race_format  = "double" if double else "single"
                flags        = SEASON_FLAGS.get(sheet_name, {})
                num_rounds   = max((rnum for _, _, rnum in race_cols), default=0)

                cur.execute(
                    """
                    INSERT INTO seasons
                        (name, display_name, score_type, race_format, sort_order,
                         champion, has_drop_round, is_multiclass)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        sheet_name, display_name, score_type, race_format, sort_order,
                        champion_raw,
                        flags.get("has_drop_round", False),
                        flags.get("is_multiclass", False),
                    ),
                )
                season_id = cur.lastrowid

                # Insert rounds
                for rnum in range(1, num_rounds + 1):
                    cur.execute(
                        "INSERT INTO rounds (season_id, round_number) VALUES (%s, %s)",
                        (season_id, rnum),
                    )

                # Insert race results row by row
                for row in rows[1:]:
                    if driver_col >= len(row):
                        continue
                    raw_driver = row[driver_col]
                    if raw_driver is None:
                        continue
                    raw_str = str(raw_driver).strip()
                    if not raw_str or raw_str.upper() == "DNP":
                        continue

                    driver_id = driver_id_map.get(raw_str)
                    if driver_id is None:
                        continue

                    for sub_type, col_i, round_num in race_cols:
                        cell_val = row[col_i] if col_i < len(row) else None
                        num, flag, starred = normalize_cell(cell_val)

                        # Skip fully empty cells
                        if num is None and flag is None:
                            continue

                        cur.execute(
                            """
                            INSERT INTO race_results
                                (season_id, driver_id, round_number, sub_type,
                                 value_numeric, value_flag, is_asterisked)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                value_numeric = VALUES(value_numeric),
                                value_flag    = VALUES(value_flag),
                                is_asterisked = VALUES(is_asterisked)
                            """,
                            (season_id, driver_id, round_num, sub_type,
                             num, flag, starred),
                        )
                        total_results += 1

                print(f"  [{sort_order+1:02d}] {sheet_name:<25} "
                      f"{num_rounds} rounds  champion={champion_raw}")

        conn.commit()
        print(f"\nImport complete.")
        print(f"  Seasons:  {len(sheet_data)}")
        print(f"  Drivers:  {len(all_raw_names)}")
        print(f"  Results:  {total_results}")

    except Exception as exc:
        conn.rollback()
        print(f"\nImport FAILED — transaction rolled back. Error: {exc}", file=sys.stderr)
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import sbr-stats.xlsx into MariaDB")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to sbr-stats.xlsx")
    args = parser.parse_args()
    run_import(args.xlsx)
